"""
Construye el reporte de RAMEC con la lógica SI / NO / SIN_DATO / NO_APLICA.

La hoja VALIDACION_PROFESIONAL ya no considera suficiente la presencia de una
caja detectada. En documentos, informa responsables, firma real por fila, fecha
de validación y una conclusión final.

Además, consolida:
- GENERAL
- VALIDACION_PAGINAS
- VALIDACION_LOGOS

Las hojas técnicas página por página se conservan ocultas para auditoría.
"""

import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extract import _norm_code, _norm_text_cmp, limpiar_titulo_control, canon_codigos


# ================== FILAS BASE ==================

def fila_estandar(path, tipo, campos, orden_ok, catalogo_ok, errores):
    return {
        "Ruta": str(path), "Archivo": path.name, "Tipo": tipo,
        "EXPEDIENTE": campos["EXPEDIENTE"], "ORIGINADOR": campos["ORIGINADOR"],
        "DIVISION": campos["DIVISION"], "FASE": campos["FASE"], "DOCUMENTO": campos["DOCUMENTO"],
        "DISCIPLINA": campos["DISCIPLINA"], "AREA": campos["AREA"], "NUMDOC": campos["NUMDOC"],
        "EXT_token": campos["EXT_token"], "EXT_real": campos["EXT_real"],
        "A_Orden_OK": "SI" if orden_ok else "NO",
        "B_Catalogo_OK": "SI" if catalogo_ok else "NO",
        "Detalle_errores": errores,
    }


def fila_comp_plano(path, expected, code, coincide):
    return {
        "Ruta": str(path), "Archivo": path.name, "Nombre_sin_ext": expected,
        "Codigo_rotulo": code,
        "Coincide": "SI" if coincide else ("NO" if code else "SIN_DATO"),
    }


def fila_comp_doc(path, expected, nd):
    coincide = (_norm_code(nd) == _norm_code(expected)) if nd else False
    return {
        "Ruta": str(path), "Archivo": path.name, "Nombre_sin_ext": expected,
        "No_Doc_caratula": nd,
        "Coincide": "SI" if coincide else ("NO" if nd else "SIN_DATO"),
    }


def fila_coherencia(path, titulo, nd, cod_doc_nombre, tipo_texto):
    return {
        "Ruta": str(path), "Archivo": path.name,
        "Titulo_detectado_caratula": titulo,
        "No_Doc_caratula": nd,
        "Cod_doc_nombre": cod_doc_nombre,
        "Tipo_detectado_caratula": tipo_texto,
        "Coherente": "SI" if (tipo_texto and tipo_texto == cod_doc_nombre) else "NO",
    }


def fila_control_cambios(path, existe, fecha_caratula, fecha_ctrl,
                         titulo, titulo_ctrl, nd, nodoc_ctrl):
    if not existe:
        fecha_ok = titulo_ok = nodoc_ok = "NO_APLICA"
        obs = "No tiene hoja de control de cambios en página 2"
    else:
        fecha_ok = "SI" if (fecha_caratula and fecha_ctrl and fecha_caratula == fecha_ctrl) else "NO"
        titulo_ok = "SI" if (_norm_text_cmp(titulo) and
                             canon_codigos(_norm_text_cmp(titulo)) ==
                             canon_codigos(limpiar_titulo_control(titulo_ctrl))) else "NO"
        nodoc_ok = "SI" if (_norm_code(nd) and _norm_code(nd) == _norm_code(nodoc_ctrl)) else "NO"
        obs_parts = []
        if not fecha_caratula:
            obs_parts.append("No se pudo leer fecha de carátula")
        if not fecha_ctrl:
            obs_parts.append("No se pudo leer fecha de control de cambios")
        if not titulo_ctrl:
            obs_parts.append("No se pudo leer título en control de cambios")
        if not nodoc_ctrl:
            obs_parts.append("No se pudo leer No. Doc. en control de cambios")
        obs = " | ".join(obs_parts)
    return {
        "Ruta": str(path), "Archivo": path.name,
        "Control_Cambios_Existe": "SI" if existe else "NO",
        "Fecha_Caratula": fecha_caratula, "Fecha_Ultimo_Cambio": fecha_ctrl,
        "Fecha_Coincide": fecha_ok,
        "Titulo_Caratula": titulo, "Titulo_Control_Cambios": limpiar_titulo_control(titulo_ctrl),
        "Titulo_Coincide": titulo_ok,
        "NoDoc_Caratula": nd, "NoDoc_Control_Cambios": nodoc_ctrl,
        "NoDoc_Coincide": nodoc_ok,
        "Observacion": obs,
    }


def _sn(valor):
    return "SI" if bool(valor) else "NO"


def fila_validacion_profesional(path, tipo, datos):
    """Construye la fila de validación profesional.

    PLANO:
        Valida presencia de responsables, validación profesional y entidades.

    DOCUMENTO:
        - Extrae Elaboró, Revisó, Aprobó 1 y Aprobó 2 (si existe).
        - Verifica una rúbrica real por cada fila.
        - Extrae la fecha de validación.
        - Compara esa fecha con carátula y última revisión.
        - Concluye SI únicamente si responsables, firmas y fecha están completos.
    """
    if tipo == "PLANO":
        responsables = datos.get("responsables", False)
        validacion = datos.get("validacion_profesional", False)
        logos = datos.get("entidades", False)
        ok = responsables and validacion and logos
        return {
            "Ruta": str(path), "Archivo": path.name, "Tipo": tipo,
            "Responsables_Plano": _sn(responsables),
            "Validacion_Profesional_Plano": _sn(validacion),
            "Logo_Entidades": _sn(logos),
            "Validacion_profesional": _sn(ok),
            "CUMPLE": "OK" if ok else "NO",
        }

    nombres = list(datos.get("nombres", []))
    firmas = list(datos.get("firmas", []))
    filas = int(datos.get("filas", 0) or 0)

    filas = max(filas, len(nombres), len(firmas))
    # Los tres roles mínimos siempre son obligatorios.
    filas = max(3, min(4, filas or 3))

    nombres = (nombres + [""] * 4)[:4]
    firmas = (firmas + [False] * 4)[:4]

    aplica_aprobo2 = filas >= 4
    requeridos = 4 if aplica_aprobo2 else 3

    responsables_completos = all(bool(nombres[i].strip()) for i in range(requeridos))
    firmas_completas = all(bool(firmas[i]) for i in range(requeridos))

    fecha_validacion = datos.get("fecha_validacion", "")
    fechas_validacion = list(datos.get("fechas_validacion", []))
    fechas_uniformes = bool(fecha_validacion) and len(set(fechas_validacion or [fecha_validacion])) == 1

    fecha_caratula = datos.get("fecha_caratula", "")
    fecha_ultima_revision = datos.get("fecha_ultima_revision", "")
    fecha_coincide = bool(
        fechas_uniformes
        and fecha_caratula
        and fecha_ultima_revision
        and fecha_validacion == fecha_caratula == fecha_ultima_revision
    )

    validacion_ok = responsables_completos and firmas_completas and fecha_coincide

    return {
        "Ruta": str(path),
        "Archivo": path.name,
        "Tipo": tipo,
        "Elaboró": nombres[0],
        "Revisó": nombres[1],
        "Aprobó_1": nombres[2],
        "Aprobó_2": nombres[3] if aplica_aprobo2 else "",
        "Firma_Elaboró": _sn(firmas[0]),
        "Firma_Revisó": _sn(firmas[1]),
        "Firma_Aprobó_1": _sn(firmas[2]),
        "Firma_Aprobó_2": _sn(firmas[3]) if aplica_aprobo2 else "NO_APLICA",
        "Fecha_validacion": fecha_validacion,
        "Responsables_completos": _sn(responsables_completos),
        "Firmas_completas": _sn(firmas_completas),
        "Fecha_coincide": _sn(fecha_coincide),
        "Validacion_profesional": _sn(validacion_ok),
        "CUMPLE": "OK" if validacion_ok else "NO",
    }


# ================== HELPERS REPORTE FINAL ==================

def _limpiar(x):
    if pd.isna(x):
        return ""
    x = str(x).replace("\n", " ")
    x = re.sub(r"\s+", " ", x).strip()
    if x.lower() in ("nan", "none", "null"):
        return ""
    return x


def _normalizar(x):
    x = _limpiar(x)
    x = unicodedata.normalize("NFKD", x)
    x = "".join(ch for ch in x if not unicodedata.combining(ch))
    x = x.upper()
    x = re.sub(r"[^A-Z0-9 ]", " ", x)
    x = re.sub(r"\s+", " ", x).strip()
    return x


def _estado_ok_no(x):
    v = _normalizar(x)
    if v in ("OK", "SI", "SÍ", "CONFORME", "CUMPLE", "TRUE"):
        return "OK"
    if v in ("NO", "OBSERVADO", "NO CUMPLE", "FALSE", "SIN DATO", "SIN_DATO"):
        return "NO"
    return ""


def _excel_safe_value(value):
    """Evita que Excel interprete textos OCR como fórmulas."""
    if isinstance(value, str) and value.strip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_df(data):
    df = pd.DataFrame(data)
    if df.empty:
        return df
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(_excel_safe_value)
    return df


def _unir_unicos(valores):
    salida = []
    for value in valores:
        value = _limpiar(value)
        if not value:
            continue
        partes = [p.strip() for p in value.split("|") if p.strip()]
        for p in partes:
            if p and p not in salida:
                salida.append(p)
    return " | ".join(salida)


def _limpiar_cip_texto(value):
    value = _limpiar(value)
    partes = [p.strip() for p in value.split("|") if p.strip()]
    salida = []
    for part in partes:
        part = re.sub(r"\.0$", "", part)
        part = re.sub(r"[^0-9]", "", part)
        if part and part not in salida:
            salida.append(part)
    return " | ".join(salida)


def _lista_paginas(df):
    if df is None or df.empty or "Página" not in df.columns:
        return ""
    pags = sorted(df["Página"].dropna().astype(int).unique().tolist())
    return ", ".join(map(str, pags))


def _responsables_por_archivo(df_validacion):
    out = {}
    if df_validacion is None or df_validacion.empty or "Archivo" not in df_validacion.columns:
        return out

    columnas = [c for c in ("Elaboró", "Revisó", "Aprobó_1", "Aprobó_2") if c in df_validacion.columns]

    for _, row in df_validacion.iterrows():
        archivo = _limpiar(row.get("Archivo", ""))
        if not archivo:
            continue
        nombres = []
        for col in columnas:
            val = _limpiar(row.get(col, ""))
            if not val:
                continue
            if _normalizar(val) in ("SI", "NO", "OK", "NO APLICA", "NO_APLICA"):
                continue
            if len(_normalizar(val).split()) >= 2 and val not in nombres:
                nombres.append(val)
        out[archivo] = " | ".join(nombres)
    return out


def _comparar_firmantes(responsables, firmantes):
    responsables = _limpiar(responsables)
    firmantes = _limpiar(firmantes)

    if not responsables and firmantes:
        return "SIN BASE", "", ""
    if responsables and not firmantes:
        return "NO", responsables, ""
    if not responsables and not firmantes:
        return "NO", "", ""

    firmantes_norm = _normalizar(firmantes)
    detectados, no_detectados = [], []

    for resp in [r.strip() for r in responsables.split("|") if r.strip()]:
        tokens = [t for t in _normalizar(resp).split() if len(t) >= 4]
        coincidencias = sum(1 for t in tokens if t in firmantes_norm)
        if coincidencias >= 2:
            detectados.append(resp)
        else:
            no_detectados.append(resp)

    if no_detectados and detectados:
        return "PARCIAL", " | ".join(no_detectados), " | ".join(detectados)
    if no_detectados:
        return "NO", " | ".join(no_detectados), ""
    return "OK", "", " | ".join(detectados)


def _limpiar_entidades(entidades, texto_ocr):
    base = f"{_limpiar(entidades)} {_limpiar(texto_ocr)}"
    base_norm = _normalizar(base)

    salida = []

    if "PROINVERSION" in base_norm:
        salida.append("PROINVERSIÓN")
    if "MTC" in base_norm or "MINISTERIO" in base_norm or "TRANSPORTES" in base_norm:
        salida.append("MTC")
    if "OSITRAN" in base_norm:
        salida.append("OSITRAN")
    if "ANILLO VIAL" in base_norm or "CONCESIONARIA" in base_norm or "AVP" in base_norm:
        salida.append("Sociedad Concesionaria Anillo Vial")

    final = []
    for e in salida:
        if e not in final:
            final.append(e)
    return " | ".join(final)


def _entidades_esperadas_ok(entidad_esperada, entidades_detectadas):
    esperadas = [e.strip() for e in _limpiar(entidad_esperada).split("|") if e.strip()]
    if not esperadas:
        esperadas = ["PROINVERSIÓN"]

    detectado = _normalizar(entidades_detectadas)

    faltantes = []
    for esp in esperadas:
        esp_norm = _normalizar(esp)

        if "PROINVERSION" in esp_norm:
            ok = "PROINVERSION" in detectado
        elif "ANILLO" in esp_norm or "CONCESIONARIA" in esp_norm:
            ok = "ANILLO VIAL" in detectado or "CONCESIONARIA" in detectado or "AVP" in detectado
        else:
            ok = esp_norm in detectado

        if not ok:
            faltantes.append(esp)

    return len(faltantes) == 0, " | ".join(faltantes)


def _resumen_validacion_paginas(df_paginas, df_validacion):
    if df_paginas is None or df_paginas.empty:
        return pd.DataFrame()

    responsables_map = _responsables_por_archivo(df_validacion)
    rows = []

    for archivo, g in df_paginas.groupby("Archivo"):
        g = g.copy()
        g["Página"] = g["Página"].astype(int)

        contenido = g[g["Tipo_página"] == "Contenido"]
        total_contenido = contenido["Página"].nunique()
        paginas_con_sello = contenido[contenido["Sello_detectado"] == "Sí"]["Página"].nunique()
        paginas_sin_sello = contenido[contenido["Sello_detectado"] != "Sí"]

        porcentaje_sello = round((paginas_con_sello / total_contenido) * 100, 2) if total_contenido else 0

        firmantes = _unir_unicos(g["Firmante_detectado"]) if "Firmante_detectado" in g.columns else ""
        cargos = _unir_unicos(g["Cargo_detectado"]) if "Cargo_detectado" in g.columns else ""
        cips = _limpiar_cip_texto(_unir_unicos(g["CIP_detectado"])) if "CIP_detectado" in g.columns else ""

        responsables = responsables_map.get(archivo, "")
        firmantes_coinciden, responsables_no_detectados, responsables_detectados = _comparar_firmantes(
            responsables, firmantes
        )

        obs = []
        if total_contenido == 0:
            obs.append("No se identificaron páginas de contenido para evaluar sellos.")
        elif paginas_con_sello == total_contenido:
            obs.append("Se detectaron sellos laterales en todas las páginas de contenido evaluadas.")
        else:
            obs.append(f"Páginas de contenido sin sello lateral detectado: {_lista_paginas(paginas_sin_sello)}.")

        if not firmantes:
            obs.append("No se identificaron firmantes en páginas de contenido.")
        if not cips:
            obs.append("No se identificaron CIP en páginas de contenido.")

        if not responsables:
            obs.append("No se logró recuperar responsables desde la hoja de control para comparación automática.")
        elif firmantes_coinciden in ("NO", "PARCIAL"):
            obs.append("Los firmantes detectados no coinciden completamente con los responsables registrados en la hoja de control.")

        cumple = "OK"
        if total_contenido == 0:
            cumple = "NO"
        if paginas_con_sello < total_contenido:
            cumple = "NO"
        if not firmantes or not cips:
            cumple = "NO"
        if firmantes_coinciden in ("NO", "PARCIAL", "SIN BASE"):
            cumple = "NO"

        rows.append({
            "Archivo": archivo,
            "Sello_detectado_en_contenido": "Sí" if paginas_con_sello > 0 else "No",
            "%_páginas_con_sello": porcentaje_sello,
            "Cantidad_páginas_con_sello": paginas_con_sello,
            "Firmantes_detectados": firmantes,
            "Cargos_detectados": cargos,
            "Responsables_hoja_de_control": responsables,
            "Responsables_detectados": responsables_detectados,
            "Responsables_no_detectados": responsables_no_detectados,
            "Firmantes_coinciden": firmantes_coinciden,
            "CIP_detectados": cips,
            "CUMPLE": cumple,
            "Observación_general": " | ".join(obs),
        })

    return pd.DataFrame(rows)


def _resumen_validacion_logos(df_paginas, entidad_esperada):
    if df_paginas is None or df_paginas.empty:
        return pd.DataFrame()

    rows = []

    for archivo, g in df_paginas.groupby("Archivo"):
        g = g.copy()
        g["Página"] = g["Página"].astype(int)

        total_paginas = g["Página"].nunique()
        paginas_con_logo = g[g["Logo_detectado_OCR"] == "Sí"]["Página"].nunique()
        paginas_sin_logo = g[g["Logo_detectado_OCR"] != "Sí"]

        porcentaje_logo = round((paginas_con_logo / total_paginas) * 100, 2) if total_paginas else 0

        entidades = _limpiar_entidades(
            _unir_unicos(g["Entidades_logo_detectadas"]) if "Entidades_logo_detectadas" in g.columns else "",
            _unir_unicos(g["Texto_logo_extraido"]) if "Texto_logo_extraido" in g.columns else "",
        )

        if not entidades and paginas_con_logo > 0:
            entidades = "Logo detectado; entidad no identificada por OCR"

        esperado_ok, faltantes = _entidades_esperadas_ok(entidad_esperada, entidades)

        obs = []
        if paginas_con_logo == total_paginas:
            obs.append("Se detectó logo en todas las páginas evaluadas.")
        else:
            obs.append(f"Páginas sin logo detectado: {_lista_paginas(paginas_sin_logo)}.")

        if esperado_ok:
            obs.append(f"Se identificó la entidad esperada: {entidad_esperada}.")
        else:
            if "MTC" in _normalizar(entidades):
                obs.append(f"Se detectó MTC, pero falta identificar: {faltantes}.")
            else:
                obs.append(f"No se identificó mediante OCR: {faltantes}.")

        cumple = "OK" if paginas_con_logo > 0 and esperado_ok else "NO"

        rows.append({
            "Archivo": archivo,
            "Logo_detectado_documento": "Sí" if paginas_con_logo > 0 else "No",
            "%_páginas_con_logo": porcentaje_logo,
            "Entidad_esperada": entidad_esperada,
            "Entidades_detectadas": entidades,
            "Concedente_OK": "OK" if esperado_ok else "NO",
            "CUMPLE": cumple,
            "Observación_general": " | ".join(obs),
        })

    return pd.DataFrame(rows)


def _estado_estandar(row):
    return "OK" if _estado_ok_no(row.get("A_Orden_OK")) == "OK" and _estado_ok_no(row.get("B_Catalogo_OK")) == "OK" else "NO"


def _estado_col(row, col):
    return _estado_ok_no(row.get(col, ""))


def _estado_control(row):
    vals = [
        _estado_ok_no(row.get("Control_Cambios_Existe", "")),
        _estado_ok_no(row.get("Fecha_Coincide", "")),
        _estado_ok_no(row.get("Titulo_Coincide", "")),
        _estado_ok_no(row.get("NoDoc_Coincide", "")),
    ]
    return "OK" if vals and all(v == "OK" for v in vals) else "NO"


def _generar_general(df_estandar, df_comp_plano, df_comp_doc, df_coherencia, df_control,
                     df_validacion, df_val_paginas, df_val_logos):
    archivos = []
    ruta_map = {}

    for df in [df_estandar, df_comp_plano, df_comp_doc, df_coherencia, df_control, df_validacion, df_val_paginas, df_val_logos]:
        if df is None or df.empty or "Archivo" not in df.columns:
            continue
        for _, row in df.iterrows():
            archivo = _limpiar(row.get("Archivo", ""))
            if archivo and archivo not in archivos:
                archivos.append(archivo)
            ruta = _limpiar(row.get("Ruta", ""))
            if archivo and ruta:
                ruta_map[archivo] = ruta

    maps = {
        "ESTANDAR NOMENCLATURA": {},
        "COMPATIBILIDAD_PLANO": {},
        "COMPATIBILIDAD_DOCUMENTO": {},
        "COHERENCIA_DOCUMENTO": {},
        "CONTROL_CAMBIOS_DOC": {},
        "VALIDACION_PROFESIONAL": {},
        "VALIDACION_PAGINAS": {},
        "VALIDACION_LOGOS": {},
    }

    if df_estandar is not None and not df_estandar.empty:
        for _, row in df_estandar.iterrows():
            maps["ESTANDAR NOMENCLATURA"][_limpiar(row.get("Archivo"))] = _estado_estandar(row)

    if df_comp_plano is not None and not df_comp_plano.empty:
        for _, row in df_comp_plano.iterrows():
            maps["COMPATIBILIDAD_PLANO"][_limpiar(row.get("Archivo"))] = _estado_col(row, "Coincide")

    if df_comp_doc is not None and not df_comp_doc.empty:
        for _, row in df_comp_doc.iterrows():
            maps["COMPATIBILIDAD_DOCUMENTO"][_limpiar(row.get("Archivo"))] = _estado_col(row, "Coincide")

    if df_coherencia is not None and not df_coherencia.empty:
        for _, row in df_coherencia.iterrows():
            maps["COHERENCIA_DOCUMENTO"][_limpiar(row.get("Archivo"))] = _estado_col(row, "Coherente")

    if df_control is not None and not df_control.empty:
        for _, row in df_control.iterrows():
            maps["CONTROL_CAMBIOS_DOC"][_limpiar(row.get("Archivo"))] = _estado_control(row)

    if df_validacion is not None and not df_validacion.empty:
        for _, row in df_validacion.iterrows():
            estado = _estado_ok_no(row.get("CUMPLE", ""))
            if not estado:
                estado = _estado_ok_no(row.get("Validacion_profesional", ""))
            maps["VALIDACION_PROFESIONAL"][_limpiar(row.get("Archivo"))] = estado

    if df_val_paginas is not None and not df_val_paginas.empty:
        for _, row in df_val_paginas.iterrows():
            maps["VALIDACION_PAGINAS"][_limpiar(row.get("Archivo"))] = _estado_ok_no(row.get("CUMPLE", ""))

    if df_val_logos is not None and not df_val_logos.empty:
        for _, row in df_val_logos.iterrows():
            maps["VALIDACION_LOGOS"][_limpiar(row.get("Archivo"))] = _estado_ok_no(row.get("CUMPLE", ""))

    rows = []
    for archivo in archivos:
        row = {
            "Ruta": ruta_map.get(archivo, ""),
            "Archivo": archivo,
            "ESTANDAR NOMENCLATURA": maps["ESTANDAR NOMENCLATURA"].get(archivo, ""),
        }

        if df_comp_plano is not None and not df_comp_plano.empty:
            row["COMPATIBILIDAD_PLANO"] = maps["COMPATIBILIDAD_PLANO"].get(archivo, "")

        if df_comp_doc is not None and not df_comp_doc.empty:
            row["COMPATIBILIDAD_DOCUMENTO"] = maps["COMPATIBILIDAD_DOCUMENTO"].get(archivo, "")

        row.update({
            "COHERENCIA_DOCUMENTO": maps["COHERENCIA_DOCUMENTO"].get(archivo, ""),
            "CONTROL_CAMBIOS_DOC": maps["CONTROL_CAMBIOS_DOC"].get(archivo, ""),
            "VALIDACION_PROFESIONAL": maps["VALIDACION_PROFESIONAL"].get(archivo, ""),
            "VALIDACION_PAGINAS": maps["VALIDACION_PAGINAS"].get(archivo, ""),
            "VALIDACION_LOGOS": maps["VALIDACION_LOGOS"].get(archivo, ""),
        })
        rows.append(row)

    return pd.DataFrame(rows)


def _aplicar_formato(salida):
    wb = load_workbook(salida)

    orden = [
        "GENERAL",
        "ESTANDAR NOMENCLATURA",
        "COMPATIBILIDAD_PLANO",
        "COMPATIBILIDAD_DOCUMENTO",
        "COHERENCIA_DOCUMENTO",
        "CONTROL_CAMBIOS_DOC",
        "VALIDACION_PROFESIONAL",
        "VALIDACION_PAGINAS",
        "VALIDACION_LOGOS",
        "VALIDACION_FINAL_PAGINAS",
        "DETALLE_SELLOS_PAGINAS",
        "DETALLE_LOGOS_PAGINAS",
    ]

    for nombre in reversed(orden):
        if nombre in wb.sheetnames:
            ws = wb[nombre]
            wb._sheets.remove(ws)
            wb._sheets.insert(0, ws)

    azul = "1F4E78"
    verde = "C6EFCE"
    rojo = "FFC7CE"
    amarillo = "FFF2CC"

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        # Evita reparación de Excel por fórmulas originadas desde OCR.
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if cell.data_type == "f":
                    cell.value = "'" + str(value)
                elif isinstance(value, str) and value.strip().startswith(("=", "+", "-", "@")):
                    cell.value = "'" + value

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=azul)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

                val = _normalizar(cell.value)
                if val in ("OK", "SI", "SÍ", "CONFORME"):
                    cell.fill = PatternFill("solid", fgColor=verde)
                elif val in ("NO", "OBSERVADO"):
                    cell.fill = PatternFill("solid", fgColor=rojo)
                elif val in ("PARCIAL", "SIN BASE", "NO APLICA", "NO_APLICA") or "REVISAR" in val:
                    cell.fill = PatternFill("solid", fgColor=amarillo)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))

            if col_letter in ("A", "B"):
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
            else:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 35 if r == 1 else 45

    for hoja in ("VALIDACION_FINAL_PAGINAS", "DETALLE_SELLOS_PAGINAS", "DETALLE_LOGOS_PAGINAS"):
        if hoja in wb.sheetnames:
            wb[hoja].sheet_state = "hidden"

    wb.save(salida)


# ================== ESCRITURA DEL REPORTE ==================

def write_report(salida, estandar, comp_plano, comp_doc, coherencia, control, validacion=None,
                 paginas=None, detalle_sellos=None, detalle_logos=None,
                 entidad_esperada="PROINVERSIÓN"):
    df_estandar = _safe_df(estandar)
    df_comp_plano = _safe_df(comp_plano)
    df_comp_doc = _safe_df(comp_doc)
    df_coherencia = _safe_df(coherencia)
    df_control = _safe_df(control)

    df_validacion = _safe_df(validacion or [])
    if not df_validacion.empty:
        preferidas = [
            "Ruta", "Archivo", "Tipo",
            "Elaboró", "Revisó", "Aprobó_1", "Aprobó_2",
            "Firma_Elaboró", "Firma_Revisó", "Firma_Aprobó_1", "Firma_Aprobó_2",
            "Fecha_validacion", "Responsables_completos", "Firmas_completas",
            "Fecha_coincide", "Validacion_profesional", "CUMPLE",
        ]
        orden = [c for c in preferidas if c in df_validacion.columns]
        orden += [c for c in df_validacion.columns if c not in orden]
        df_validacion = df_validacion[orden]

    df_paginas = _safe_df(paginas or [])
    df_detalle_sellos = _safe_df(detalle_sellos or [])
    df_detalle_logos = _safe_df(detalle_logos or [])

    df_validacion_paginas = _safe_df(_resumen_validacion_paginas(df_paginas, df_validacion))
    df_validacion_logos = _safe_df(_resumen_validacion_logos(df_paginas, entidad_esperada))
    df_general = _safe_df(_generar_general(
        df_estandar,
        df_comp_plano,
        df_comp_doc,
        df_coherencia,
        df_control,
        df_validacion,
        df_validacion_paginas,
        df_validacion_logos,
    ))

    with pd.ExcelWriter(salida, engine="openpyxl") as xw:
        if not df_general.empty:
            df_general.to_excel(xw, sheet_name="GENERAL", index=False)

        df_estandar.to_excel(xw, sheet_name="ESTANDAR NOMENCLATURA", index=False)

        if not df_comp_plano.empty:
            df_comp_plano.to_excel(xw, sheet_name="COMPATIBILIDAD_PLANO", index=False)
        if not df_comp_doc.empty:
            df_comp_doc.to_excel(xw, sheet_name="COMPATIBILIDAD_DOCUMENTO", index=False)
        if not df_coherencia.empty:
            df_coherencia.to_excel(xw, sheet_name="COHERENCIA_DOCUMENTO", index=False)
        if not df_control.empty:
            df_control.to_excel(xw, sheet_name="CONTROL_CAMBIOS_DOC", index=False)
        if not df_validacion.empty:
            df_validacion.to_excel(xw, sheet_name="VALIDACION_PROFESIONAL", index=False)
        if not df_validacion_paginas.empty:
            df_validacion_paginas.to_excel(xw, sheet_name="VALIDACION_PAGINAS", index=False)
        if not df_validacion_logos.empty:
            df_validacion_logos.to_excel(xw, sheet_name="VALIDACION_LOGOS", index=False)

        if not df_paginas.empty:
            df_paginas.to_excel(xw, sheet_name="VALIDACION_FINAL_PAGINAS", index=False)
        if not df_detalle_sellos.empty:
            df_detalle_sellos.to_excel(xw, sheet_name="DETALLE_SELLOS_PAGINAS", index=False)
        if not df_detalle_logos.empty:
            df_detalle_logos.to_excel(xw, sheet_name="DETALLE_LOGOS_PAGINAS", index=False)

    _aplicar_formato(salida)
